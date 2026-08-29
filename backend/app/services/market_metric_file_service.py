import csv
import hashlib
import io
import json
import os
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import ValidationError

from app.modules.market_intelligence.schemas import (
    DirectMarketMetricInput,
    MarketMetricStoredFile,
)


class MarketMetricFileError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class ParsedMarketMetricFile:
    stored_file: MarketMetricStoredFile
    metrics: list[DirectMarketMetricInput]


class MarketMetricFileService:
    """安全解析并原子保存运营人员上传的宏观指标文件。"""

    reference_prefix = "market-metric-upload://"
    supported_content_types = {
        ".json": {"application/json", "text/json", "application/octet-stream"},
        ".csv": {
            "text/csv",
            "application/csv",
            "application/vnd.ms-excel",
            "text/plain",
            "application/octet-stream",
        },
        ".xlsx": {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/zip",
            "application/octet-stream",
        },
    }

    def __init__(
        self,
        *,
        storage_root: Path,
        max_bytes: int,
        max_uncompressed_bytes: int,
        max_rows: int,
    ) -> None:
        self.storage_root = storage_root.resolve()
        self.max_bytes = max_bytes
        self.max_uncompressed_bytes = max_uncompressed_bytes
        self.max_rows = max_rows

    def ingest(
        self,
        *,
        tenant_id: str,
        filename: str | None,
        content_type: str | None,
        content: bytes,
    ) -> ParsedMarketMetricFile:
        extension = Path(filename or "").suffix.casefold()
        self._validate_upload(extension, content_type, content)
        metrics = self._parse(extension, content)
        stored_file = self._store(tenant_id, extension, content)
        return ParsedMarketMetricFile(stored_file=stored_file, metrics=metrics)

    def discard(self, stored_file: MarketMetricStoredFile) -> None:
        if not stored_file.file_ref.startswith(self.reference_prefix):
            return
        relative = stored_file.file_ref.removeprefix(self.reference_prefix)
        target = (self.storage_root / Path(relative)).resolve()
        if self.storage_root not in target.parents:
            return
        try:
            target.unlink(missing_ok=True)
        except OSError:
            return

    def _validate_upload(
        self,
        extension: str,
        content_type: str | None,
        content: bytes,
    ) -> None:
        if extension not in self.supported_content_types:
            raise MarketMetricFileError(
                "UNSUPPORTED_FILE_TYPE",
                "仅支持 JSON、CSV 和 XLSX 格式。",
            )
        media_type = (
            (content_type or "application/octet-stream")
            .split(";", 1)[0]
            .strip()
        )
        if media_type not in self.supported_content_types[extension]:
            raise MarketMetricFileError(
                "CONTENT_TYPE_MISMATCH",
                "文件扩展名与 Content-Type 不匹配。",
            )
        if not content:
            raise MarketMetricFileError("EMPTY_UPLOAD_FILE", "上传文件不能为空。")
        if len(content) > self.max_bytes:
            raise MarketMetricFileError(
                "UPLOAD_FILE_TOO_LARGE",
                f"上传文件不能超过 {self.max_bytes} 字节。",
            )
        if extension == ".xlsx":
            self._validate_xlsx_archive(content)

    def _validate_xlsx_archive(self, content: bytes) -> None:
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                expanded_size = sum(item.file_size for item in archive.infolist())
        except (OSError, zipfile.BadZipFile) as exc:
            raise MarketMetricFileError(
                "INVALID_XLSX_FILE",
                "XLSX 文件结构无效。",
            ) from exc
        if expanded_size > self.max_uncompressed_bytes:
            raise MarketMetricFileError(
                "XLSX_EXPANDED_SIZE_EXCEEDED",
                "XLSX 解压后内容超过安全限制。",
            )

    def _parse(self, extension: str, content: bytes) -> list[DirectMarketMetricInput]:
        if extension == ".json":
            rows = self._json_rows(content)
        elif extension == ".csv":
            rows = self._csv_rows(content)
        else:
            rows = self._xlsx_rows(content)
        if not rows:
            raise MarketMetricFileError("MARKET_METRIC_FILE_EMPTY", "文件中没有指标数据。")
        if len(rows) > self.max_rows:
            raise MarketMetricFileError(
                "MARKET_METRIC_ROW_LIMIT_EXCEEDED",
                f"单次最多上传 {self.max_rows} 条指标。",
            )
        metrics: list[DirectMarketMetricInput] = []
        for index, row in enumerate(rows, start=1):
            try:
                metrics.append(DirectMarketMetricInput.model_validate(row))
            except ValidationError as exc:
                message = "; ".join(
                    f"{'.'.join(map(str, item['loc']))}: {item['msg']}"
                    for item in exc.errors(include_url=False)
                )
                raise MarketMetricFileError(
                    "MARKET_METRIC_ROW_INVALID",
                    f"第 {index} 条指标无效：{message}",
                ) from exc
        codes = [item.metric_code for item in metrics]
        if len(codes) != len(set(codes)):
            raise MarketMetricFileError(
                "DUPLICATE_MARKET_METRIC_CODE",
                "同一文件中的 metric_code 不能重复。",
            )
        return metrics

    @staticmethod
    def _json_rows(content: bytes) -> list[dict]:
        try:
            payload = json.loads(content.decode("utf-8-sig"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise MarketMetricFileError(
                "INVALID_JSON_FILE",
                "JSON 文件必须使用 UTF-8 编码并包含合法 JSON。",
            ) from exc
        rows = payload.get("metrics") if isinstance(payload, dict) else payload
        if not isinstance(rows, list) or any(not isinstance(item, dict) for item in rows):
            raise MarketMetricFileError(
                "INVALID_JSON_STRUCTURE",
                "JSON 内容应为指标数组或包含 metrics 数组的对象。",
            )
        return rows

    def _csv_rows(self, content: bytes) -> list[dict]:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise MarketMetricFileError(
                "INVALID_CSV_ENCODING",
                "CSV 文件必须使用 UTF-8 编码。",
            ) from exc
        reader = csv.DictReader(io.StringIO(text, newline=""))
        if not reader.fieldnames:
            raise MarketMetricFileError("CSV_HEADER_MISSING", "CSV 文件缺少表头。")
        headers = self._headers(reader.fieldnames)
        rows: list[dict] = []
        for row in reader:
            if None in row:
                raise MarketMetricFileError(
                    "CSV_COLUMN_COUNT_MISMATCH",
                    "CSV 数据列数与表头不一致。",
                )
            if not any(value not in (None, "") for value in row.values()):
                continue
            rows.append(self._clean_row(dict(zip(headers, row.values()))))
            if len(rows) > self.max_rows:
                break
        return rows

    def _xlsx_rows(self, content: bytes) -> list[dict]:
        workbook = None
        try:
            workbook = load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
            worksheet = workbook.active
            iterator = worksheet.iter_rows(values_only=True)
            first = next(iterator, None)
            if first is None:
                return []
            headers = self._headers(first)
            rows: list[dict] = []
            for values in iterator:
                if not any(value not in (None, "") for value in values):
                    continue
                if len(values) > len(headers) and any(
                    value not in (None, "") for value in values[len(headers):]
                ):
                    raise MarketMetricFileError(
                        "XLSX_COLUMN_COUNT_MISMATCH",
                        "XLSX 数据列数与表头不一致。",
                    )
                rows.append(self._clean_row(dict(zip(headers, values))))
                if len(rows) > self.max_rows:
                    break
            return rows
        except MarketMetricFileError:
            raise
        except Exception as exc:
            raise MarketMetricFileError(
                "INVALID_XLSX_FILE",
                "无法读取 XLSX 文件内容。",
            ) from exc
        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def _headers(values) -> list[str]:
        headers = [str(value or "").strip().casefold() for value in values]
        if any(not value for value in headers) or len(headers) != len(set(headers)):
            raise MarketMetricFileError(
                "INVALID_FILE_HEADER",
                "文件表头不能为空或重复。",
            )
        return headers

    @staticmethod
    def _clean_row(row: dict) -> dict:
        return {
            key: value.strip() if isinstance(value, str) else value
            for key, value in row.items()
            if value not in (None, "")
        }

    def _store(
        self,
        tenant_id: str,
        extension: str,
        content: bytes,
    ) -> MarketMetricStoredFile:
        digest = hashlib.sha256(content).hexdigest()
        tenant_segment = hashlib.sha256(tenant_id.encode("utf-8")).hexdigest()[:24]
        relative = Path(tenant_segment) / f"{uuid4().hex}-{digest[:16]}{extension}"
        target = (self.storage_root / relative).resolve()
        if self.storage_root not in target.parents:
            raise MarketMetricFileError("UNSAFE_STORAGE_PATH", "上传文件存储路径无效。")
        temporary_name: str | None = None
        try:
            target.parent.mkdir(mode=0o700, parents=True, exist_ok=True)
            descriptor, temporary_name = tempfile.mkstemp(
                prefix="market-metric-",
                suffix=".tmp",
                dir=target.parent,
            )
            with os.fdopen(descriptor, "wb") as temporary:
                temporary.write(content)
                temporary.flush()
                os.fsync(temporary.fileno())
            os.replace(temporary_name, target)
        except OSError as exc:
            if temporary_name is not None:
                try:
                    Path(temporary_name).unlink(missing_ok=True)
                except OSError:
                    pass
            raise MarketMetricFileError(
                "UPLOAD_FILE_STORAGE_FAILED",
                "上传文件保存失败。",
            ) from exc
        return MarketMetricStoredFile(
            file_ref=f"{self.reference_prefix}{relative.as_posix()}",
            sha256=digest,
        )


__all__ = [
    "MarketMetricFileError",
    "MarketMetricFileService",
    "ParsedMarketMetricFile",
]
